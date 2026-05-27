from __future__ import annotations

import argparse
import os
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.outsource_attendance import (  # noqa: E402
    DEFAULT_MONGODB_DATABASE,
    MONGODB_DATABASE_ENV_VAR,
    MONGODB_URI_ENV_VAR,
    AttendanceService,
    MongoAttendanceService,
    _clean_department,
    _clean_gender,
    _clean_mobile,
    _clean_name,
)


HEADER_ALIASES = {
    "name": {"name"},
    "mobile": {"mobile", "mobilenumber", "phonenumber", "phone"},
    "gender": {"fm", "gender", "malefemale"},
    "department": {"noticecalling", "department", "noticecallingdepartment"},
}
CREATED_BY = "OUTSOURCE.xlsx import"


def _header_key(value: Any) -> str:
    return "".join(ch for ch in str(value or "").casefold() if ch.isalnum())


def _load_streamlit_secrets() -> dict[str, Any]:
    secrets_path = ROOT / ".streamlit" / "secrets.toml"
    if not secrets_path.exists():
        return {}
    try:
        import tomllib
    except ModuleNotFoundError:
        return {}

    with secrets_path.open("rb") as handle:
        return tomllib.load(handle)


def read_outsource_users(workbook_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook.active
    header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_indexes: dict[str, int] = {}

    for index, header in enumerate(header_values):
        key = _header_key(header)
        for field, aliases in HEADER_ALIASES.items():
            if key in aliases:
                header_indexes[field] = index

    missing = sorted(set(HEADER_ALIASES) - set(header_indexes))
    if missing:
        raise ValueError(f"Missing required workbook columns: {', '.join(missing)}")

    records: list[dict[str, str]] = []
    errors: list[str] = []
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in row):
            continue

        name = _clean_name(row[header_indexes["name"]])
        mobile = _clean_mobile(row[header_indexes["mobile"]])
        gender = _clean_gender(row[header_indexes["gender"]])
        department = _clean_department(row[header_indexes["department"]])

        if not name:
            errors.append(f"row {row_number}: missing name")
        if not mobile:
            errors.append(f"row {row_number}: missing mobile number")
        if gender not in {"Female", "Male"}:
            errors.append(f"row {row_number}: gender must be M/F, Male/Female")
        if department not in {"Notice", "Calling"}:
            errors.append(f"row {row_number}: department must be N/C, Notice/Calling")

        records.append(
            {
                "name": name,
                "mobile": mobile,
                "gender": gender,
                "department": department,
                "joined_date": "",
                "study": "",
                "details": "",
            }
        )

    for field, values in {
        "mobile": [record["mobile"] for record in records],
        "name": [record["name"].casefold() for record in records],
    }.items():
        duplicates = [value for value, count in Counter(values).items() if value and count > 1]
        if duplicates:
            errors.append(f"duplicate {field} values: {', '.join(duplicates)}")

    if errors:
        raise ValueError("\n".join(errors))

    return records


def replace_service_users(service: AttendanceService, records: list[dict[str, str]]) -> int:
    service.reset_all_data()
    for record in records:
        service.add_user(
            name=record["name"],
            role="outsource",
            mobile=record["mobile"],
            designation=record["department"],
            joined_date=record["joined_date"],
            details=record["details"],
            gender=record["gender"],
            department=record["department"],
            study=record["study"],
            created_by=CREATED_BY,
        )
    return len(records)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Replace attendance data with outsource users from OUTSOURCE.xlsx."
    )
    parser.add_argument("--workbook", default=str(ROOT / "OUTSOURCE.xlsx"))
    parser.add_argument("--local-db", default=None)
    parser.add_argument("--skip-local", action="store_true")
    parser.add_argument("--skip-mongodb", action="store_true")
    parser.add_argument("--mongodb-uri", default=None)
    parser.add_argument("--mongodb-database", default=None)
    parser.add_argument("--dry-run", action="store_true")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    workbook_path = Path(args.workbook)
    if not workbook_path.is_absolute():
        workbook_path = ROOT / workbook_path

    records = read_outsource_users(workbook_path)
    gender_counts = Counter(record["gender"] for record in records)
    department_counts = Counter(record["department"] for record in records)
    print(
        "Workbook ready: "
        f"{len(records)} users, "
        f"Male={gender_counts.get('Male', 0)}, Female={gender_counts.get('Female', 0)}, "
        f"Notice={department_counts.get('Notice', 0)}, Calling={department_counts.get('Calling', 0)}"
    )

    if args.dry_run:
        return 0

    if not args.skip_local:
        local_service = AttendanceService(args.local_db)
        loaded = replace_service_users(local_service, records)
        print(f"Local SQLite replaced: {loaded} outsource users")

    if not args.skip_mongodb:
        secrets = _load_streamlit_secrets()
        mongodb_uri = (
            args.mongodb_uri
            or os.environ.get(MONGODB_URI_ENV_VAR)
            or str(secrets.get(MONGODB_URI_ENV_VAR, "") or "")
        )
        mongodb_database = (
            args.mongodb_database
            or os.environ.get(MONGODB_DATABASE_ENV_VAR)
            or str(secrets.get(MONGODB_DATABASE_ENV_VAR, "") or "")
            or DEFAULT_MONGODB_DATABASE
        )
        if not mongodb_uri:
            print("MongoDB skipped: MONGODB_URI was not provided")
        else:
            mongo_service = MongoAttendanceService(mongodb_uri, mongodb_database)
            loaded = replace_service_users(mongo_service, records)
            print(f"MongoDB Atlas replaced: {loaded} outsource users in {mongodb_database}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
